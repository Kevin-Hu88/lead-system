# -*- coding: utf-8 -*-

import os, re, time, threading, requests

from datetime import datetime

from flask import Flask, render_template, request, jsonify

from flask_cors import CORS

from loguru import logger

from config import settings

from crm.models import db, Lead, Message, TaskLog

from crm.database import init_db, get_dashboard_stats, update_lead_status, update_lead_score, record_message, add_lead, get_human_handoff_leads



harvest_state = {

    'running': False, 'pct': 0, 'current': '', 'added': 0,

    'with_phone': 0, 'skipped': 0, 'skip_phone': 0, 'skip_name': 0, 'api_calls': 0, 'log_lines': ''

}



project_harvest_state = {

    'running': False, 'pct': 0, 'current': '', 'added': 0,

    'scanned': 0, 'source': '', 'log_lines': ''

}



def create_app():

    app = Flask(__name__,

        template_folder=os.path.join(os.path.dirname(__file__), 'templates'),

        static_folder=os.path.join(os.path.dirname(__file__), 'static'))

    app.config['SECRET_KEY'] = settings.SECRET_KEY

    app.config['SQLALCHEMY_DATABASE_URI'] = settings.SQLALCHEMY_DATABASE_URI

    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

    CORS(app)

    init_db(app)

    return app



app = create_app()

# >>> 安全兜底开始
import tempfile
from pathlib import Path as _Path
from flask import request as _req


def _atomic_write_text(path, content, encoding='utf-8', syntax_check: bool = True):
    path = _Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile('w', encoding=encoding, delete=False, dir=str(path.parent), suffix='.tmp') as tf:
            tf.write(content)
            tmp_path = _Path(tf.name)
        tmp_path.replace(path)
    finally:
        if tmp_path and tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError:
                pass

    if syntax_check and path.suffix == '.py':
        try:
            compile(path.read_text(encoding=encoding, errors='ignore'), str(path), 'exec')
        except SyntaxError:
            pass

    return path


@app.errorhandler(Exception)
def _global_exception_handler(exc):
    try:
        if _req.path.startswith('/api'):
            return jsonify({'ok': False, 'error': str(exc)}), 500
    except Exception:
        pass
    raise exc
# <<< 安全兜底结束

def _ok(data, status=200):
    return jsonify({"ok": True, "data": data}), status


def _fail(message, status=400):
    return jsonify({"ok": False, "error": message}), status




# Auto-trigger competitor harvest if data is empty

def _auto_competitor_check():

    import time, importlib

    time.sleep(10)

    try:

        import config.competitors as comp_mod

        importlib.reload(comp_mod)

        COMPETITORS = dict(comp_mod.COMPETITORS)

        total = sum(len(v) for v in COMPETITORS.values())

        if total < 10:

            logger.info("[AutoCompetitor] Starting auto harvest...")

            from lead_harvester.competitor_harvester import CompetitorHarvester

            with app.app_context():

                harvester = CompetitorHarvester()

                saved_total = 0

                for region in list(COMPETITORS.keys()):

                    try:

                        companies = harvester.harvest_region(region)

                        if companies:

                            if region not in COMPETITORS:

                                COMPETITORS[region] = []

                            for c in companies:

                                if not any(e.get("name") == c["name"] for e in COMPETITORS[region]):

                                    COMPETITORS[region].append(c)

                                    saved_total += 1

                            _save_competitors(COMPETITORS)

                            logger.info(f"[AutoCompetitor] {region}: +{len(companies)}, saved. Total new: {saved_total}")

                        time.sleep(2)

                    except Exception as e:

                        logger.debug(f"[AutoCompetitor] {region} failed: {e}")

                logger.info(f"[AutoCompetitor] Done: +{saved_total} new companies")

    except Exception as e:

        logger.debug(f"[AutoCompetitor] Skipped: {e}")



import threading

threading.Thread(target=_auto_competitor_check, daemon=True).start()



def extract_phone(text):

    if not text: return ''

    m = re.search(r'1[3-9]\d{9}', text)

    if m: return m.group()

    m = re.search(r'0\d{2,3}[-]?\d{7,8}', text)

    if m: return m.group()

    return ''



def _do_harvest(app_ctx, areas, targets, categories=None):

    global harvest_state

    log_buf = []

    try:

        harvest_state.update({'running': True, 'pct': 0, 'current': 'Starting...',

            'added': 0, 'with_phone': 0, 'skipped': 0, 'skip_phone': 0, 'skip_name': 0, 'api_calls': 0, 'log_lines': ''})

        ak = settings.BAIDU_MAP_AK

        api_url = 'https://api.map.baidu.com/place/v2/search'

        combos = [(a, t) for a in areas for t in targets]

        total = len(combos)

        done_count = [0]

        from concurrent.futures import ThreadPoolExecutor, as_completed



        def harvest_one(combo):

            area, t = combo

            keyword = t['keyword']

            customer_type = t['customer_type']

            local_added = 0

            local_skipped = 0

            local_phone = 0

            local_api = 0

            new_leads = []

            for page in range(5):

                try:

                    params = {'query': keyword, 'region': area, 'ak': ak,

                        'output': 'json', 'scope': 2, 'page_size': 20, 'page_num': page}

                    r = requests.get(api_url, params=params, timeout=10)

                    local_api += 1

                    data = r.json()

                    if data.get('status') != 0: break

                    pois = data.get('results', [])

                    if not pois: break

                    for poi in pois:

                        name = poi.get('name', '').strip()

                        if not name: continue

                        addr = poi.get('address', '').strip()

                        phone = extract_phone(poi.get('telephone', ''))

                        if not phone:

                            phone = extract_phone(poi.get('detail_info', {}).get('telephone', ''))

                        new_leads.append({

                            'name': name[:100], 'phone': phone, 'address': addr[:300],

                            'source': 'baidu_map', 'area': area,

                            'customer_type': customer_type, 'product_interest': keyword,

                        })

                        if phone: local_phone += 1

                    if len(pois) < 20: break

                except Exception:

                    break

            return area, keyword, new_leads, local_phone, local_api



        with ThreadPoolExecutor(max_workers=8) as executor:

            futures = {executor.submit(harvest_one, c): c for c in combos}

            for future in as_completed(futures):

                try:

                    area, keyword, new_leads, local_phone, local_api = future.result()

                    harvest_state['api_calls'] += local_api

                    area_added = 0

                    area_skip_phone = 0

                    area_skip_name = 0

                    with app_ctx.app_context():

                        from crm.scoring import score_lead

                        existing_phones = set(r[0] for r in db.session.query(Lead.phone).filter(Lead.phone.isnot(None), Lead.phone != '').all())

                        existing_names = set(r[0] for r in db.session.query(Lead.name).all())

                        for ld in new_leads:

                            if ld['phone'] and ld['phone'] in existing_phones:

                                area_skip_phone += 1; continue

                            if ld['name'] in existing_names:

                                area_skip_name += 1; continue

                            lead = Lead(**ld)

                            score_lead(lead)

                            db.session.add(lead)

                            area_added += 1

                            if ld['phone']: existing_phones.add(ld['phone'])

                            existing_names.add(ld['name'])

                        db.session.commit()

                    harvest_state['added'] += area_added

                    harvest_state['skipped'] += area_skip_phone + area_skip_name

                    harvest_state['skip_phone'] += area_skip_phone

                    harvest_state['skip_name'] += area_skip_name

                    harvest_state['with_phone'] += local_phone

                    done_count[0] += 1

                    pct = int(done_count[0] / total * 100)

                    harvest_state['pct'] = pct

                    harvest_state['current'] = f'[{pct}%] {area}/{keyword}: +{area_added}'

                    skip_info = ''

                    if area_skip_phone > 0 or area_skip_name > 0:

                        skip_info = f' (skip: phone={area_skip_phone}, name={area_skip_name})'

                    log_buf.append(f'[{pct}%] {area}/{keyword}: +{area_added}{skip_info}')

                    harvest_state['log_lines'] = '\n'.join(log_buf[-30:])

                except Exception as e:

                    log_buf.append(f'  Error: {e}')

                    harvest_state['log_lines'] = '\n'.join(log_buf[-30:])

    except Exception as e:

        log_buf.append(f'Harvest error: {e}')

        harvest_state['log_lines'] = '\n'.join(log_buf[-30:])

        logger.error(f'Harvest failed: {e}')

    finally:

        harvest_state['running'] = False

        harvest_state['pct'] = 100

        harvest_state['current'] = 'Done'



def _do_project_harvest(app_ctx, source, max_pages=200, area_filter='wuhan'):

    global project_harvest_state

    log_buf = []

    try:

        project_harvest_state.update({

            'running': True, 'pct': 0, 'current': 'Starting...',

            'added': 0, 'scanned': 0, 'source': source, 'log_lines': ''

        })

        with app_ctx.app_context():

            if source == 'hbcic':

                from lead_harvester.hbcic_harvester import HbciCHarvester

                harvester = HbciCHarvester()

                def progress_cb(added, scanned, total):

                    project_harvest_state['added'] = added

                    project_harvest_state['scanned'] = scanned

                    pct = min(99, int(scanned / (max_pages * 20) * 100))

                    project_harvest_state['pct'] = pct

                    project_harvest_state['current'] = f'[{pct}%] scanned {scanned}, +{added}'

                    log_buf.append(f'[+{added}] scanned {scanned}')

                    project_harvest_state['log_lines'] = '\n'.join(log_buf[-30:])

                harvester.progress_cb = progress_cb

                count = harvester.harvest(max_pages=max_pages, area_filter=area_filter)

                project_harvest_state['added'] = count

            elif source == 'spfxm':

                from lead_harvester.spfxm_harvester import SpfxmHarvester

                harvester = SpfxmHarvester()

                count = harvester.harvest()

                project_harvest_state['added'] = count

            elif source == 'housing_bureau':

                from lead_harvester.housing_bureau_harvester import HousingBureauHarvester

                harvester = HousingBureauHarvester()

                count = harvester.harvest()

                project_harvest_state['added'] = count

            elif source == 'project':

                from lead_harvester.project_harvester import ProjectHarvester

                harvester = ProjectHarvester()

                count = harvester.harvest()

                project_harvest_state['added'] = count

            else:

                log_buf.append(f'Unknown source: {source}')



        log_buf.append(f'Done! +{project_harvest_state["added"]} project leads')

        project_harvest_state['log_lines'] = '\n'.join(log_buf[-30:])

    except Exception as e:

        log_buf.append(f'Error: {e}')

        project_harvest_state['log_lines'] = '\n'.join(log_buf[-30:])

        logger.error(f'Project harvest failed: {e}')

    finally:

        project_harvest_state['running'] = False

        project_harvest_state['pct'] = 100

        project_harvest_state['current'] = 'Done'



# ===================== Pages =====================

@app.route('/')

def index():

    stats = get_dashboard_stats()

    handoff = get_human_handoff_leads()

    recent = Lead.query.order_by(Lead.created_at.desc()).limit(20).all()

    msgs = Message.query.order_by(Message.created_at.desc()).limit(20).all()

    s_count = Lead.query.filter_by(lead_level='S').count()

    a_count = Lead.query.filter_by(lead_level='A').count()

    return render_template('index.html', stats=stats, handoff_leads=handoff,

        recent_leads=recent, recent_messages=msgs, s_count=s_count, a_count=a_count)



@app.route('/leads')

def leads_page():

    status = request.args.get('status', '')

    source = request.args.get('source', '')

    business_category = request.args.get('business_category', '')

    customer_type = request.args.get('customer_type', '')

    area = request.args.get('area', '')

    has_phone = request.args.get('has_phone', '')

    lead_level = request.args.get('lead_level', '')

    sort = request.args.get('sort', 'newest')

    page = request.args.get('page', 1, type=int)

    q = Lead.query

    if status: q = q.filter_by(status=status)

    if source: q = q.filter_by(source=source)

    if business_category: q = q.filter_by(business_category=business_category)

    if customer_type: q = q.filter_by(customer_type=customer_type)

    if area: q = q.filter(Lead.area.like(f'%{area}%'))

    if has_phone == 'yes': q = q.filter(Lead.phone.isnot(None), Lead.phone != '')

    elif has_phone == 'no': q = q.filter((Lead.phone.is_(None)) | (Lead.phone == ''))

    if lead_level: q = q.filter_by(lead_level=lead_level)

    # Date filters
    created_today = request.args.get('created_today', '')
    created_week = request.args.get('created_week', '')
    if created_today:
        today = datetime.now().date()
        q = q.filter(Lead.created_at >= datetime.combine(today, datetime.min.time()))
    if created_week:
        week_ago = datetime.now().date() - timedelta(days=7)
        q = q.filter(Lead.created_at >= datetime.combine(week_ago, datetime.min.time()))

    # Sleep status filter
    sleep_status = request.args.get('sleep_status', '')
    if sleep_status:
        q = q.filter_by(sleep_status=int(sleep_status))

    bid_min_budget = request.args.get('bid_min_budget', '')

    bid_max_budget = request.args.get('bid_max_budget', '')

    bid_deadline_from = request.args.get('bid_deadline_from', '')

    bid_deadline_to = request.args.get('bid_deadline_to', '')

    if bid_min_budget:

        try: q = q.filter(Lead.bid_budget >= float(bid_min_budget))

        except: pass

    if bid_max_budget:

        try: q = q.filter(Lead.bid_budget <= float(bid_max_budget))

        except: pass

    if sort == 'score': q = q.order_by(Lead.total_score.desc())

    elif sort == 'phone': q = q.order_by(Lead.phone.isnot(None).desc(), Lead.created_at.desc())

    elif sort == 'level': q = q.order_by(Lead.lead_level.asc(), Lead.total_score.desc())

    else: q = q.order_by(Lead.created_at.desc())

    pagination = q.paginate(page=page, per_page=50, error_out=False)

    all_sources = [r[0] for r in db.session.query(Lead.source).distinct().all() if r[0]]

    all_types = [r[0] for r in db.session.query(Lead.customer_type).distinct().all() if r[0]]

    all_areas = sorted(set(r[0] for r in db.session.query(Lead.area).distinct().all() if r[0] and ('武汉' in r[0] or '湖北' not in r[0])))

    level_counts = {}

    for lvl in ['S', 'A', 'B', 'C']:

        lcq = Lead.query.filter_by(lead_level=lvl)

        if has_phone == 'yes': lcq = lcq.filter(Lead.phone.isnot(None), Lead.phone != '')

        level_counts[lvl] = lcq.count()

    cat_counts = {'all': Lead.query.count()}

    for cat in ['膜结构', '玻璃遮阳棚', '光伏车棚']:

        cat_counts[cat] = Lead.query.filter_by(business_category=cat).count()

    return render_template('leads.html', pagination=pagination, leads=pagination.items,

        status=status, source=source, customer_type=customer_type, area=area,

        has_phone=has_phone, lead_level=lead_level, sort=sort, business_category=business_category,

        all_sources=sorted(all_sources), all_types=sorted(all_types), all_areas=all_areas,

        level_counts=level_counts, cat_counts=cat_counts)



@app.route('/lead/<int:lid>')

def lead_detail(lid):

    lead = Lead.query.get_or_404(lid)

    messages = Message.query.filter_by(lead_id=lid).order_by(Message.created_at.desc()).all()

    return render_template('lead_detail.html', lead=lead, messages=messages)



@app.route('/harvest')

def harvest_page():

    return render_template('harvest.html')



# ===================== APIs =====================

@app.route('/api/stats')

def api_stats():

    stats = get_dashboard_stats()

    stats['level_counts'] = {

        'S': Lead.query.filter_by(lead_level='S').count(),

        'A': Lead.query.filter_by(lead_level='A').count(),

        'B': Lead.query.filter_by(lead_level='B').count(),

        'C': Lead.query.filter_by(lead_level='C').count(),

    }

    stats['opt_out_count'] = Lead.query.filter_by(is_opt_out=True).count()

    stats['sleep_count'] = Lead.query.filter_by(sleep_status=1).count()

    return _ok(stats)



@app.route('/api/heatmap')

def api_heatmap():

    rows = db.session.query(Lead.area, db.func.count(Lead.id)).group_by(Lead.area).all()

    result = []

    for area, count in rows:

        if not area: continue

        phone_count = Lead.query.filter(Lead.area == area, Lead.phone.isnot(None), Lead.phone != '').count()

        won = Lead.query.filter(Lead.area == area, Lead.status == 'closed_won').count()

        s_count = Lead.query.filter(Lead.area == area, Lead.lead_level == 'S').count()

        a_count = Lead.query.filter(Lead.area == area, Lead.lead_level == 'A').count()

        result.append({'area': area, 'total': count, 'with_phone': phone_count,

            'closed_won': won, 's_count': s_count, 'a_count': a_count})

    result.sort(key=lambda x: -x['total'])

    return _ok(result)



@app.route('/api/stats/by_type')

def api_stats_by_type():

    from sqlalchemy import func

    rows = db.session.query(

        Lead.customer_type, func.count(Lead.id),

        func.sum(db.case((Lead.phone.isnot(None) & (Lead.phone != ''), 1), else_=0)),

        func.sum(db.case((Lead.lead_level == 'S', 1), else_=0)),

        func.sum(db.case((Lead.lead_level == 'A', 1), else_=0)),

    ).group_by(Lead.customer_type).all()

    result = []

    for ctype, total, with_phone, s_count, a_count in rows:

        if not ctype: continue

        result.append({'type': ctype, 'total': total, 'with_phone': with_phone or 0,

            's_count': s_count or 0, 'a_count': a_count or 0})

    result.sort(key=lambda x: -x['total'])

    return _ok(result)



@app.route('/api/stats/by_area')

def api_stats_by_area():

    from sqlalchemy import func

    rows = db.session.query(

        Lead.area, func.count(Lead.id),

        func.sum(db.case((Lead.phone.isnot(None) & (Lead.phone != ''), 1), else_=0)),

    ).group_by(Lead.area).all()

    result = []

    for area, total, with_phone in rows:

        if not area: continue

        result.append({'area': area, 'total': total, 'with_phone': with_phone or 0})

    result.sort(key=lambda x: -x['total'])

    return _ok(result)



@app.route('/api/calc_score', methods=['POST'])

def api_calc_score():

    from crm.scoring import score_all_leads

    result = score_all_leads(app)

    return _ok(result)



@app.route('/api/sleep_batch', methods=['POST'])

def api_sleep_batch():

    from crm.scoring import mark_sleep_leads

    count = mark_sleep_leads(app)

    return _ok({'sleep_count': count})



@app.route('/api/optout', methods=['POST'])

def api_optout():

    data = request.get_json()

    phone = data.get('phone', '')

    if not phone: return _fail('Need phone')

    from crm.scoring import opt_out_lead

    count = opt_out_lead(phone)

    return _ok({'opted_out': count})



@app.route('/api/pop_alert')

def api_pop_alert():

    s_leads = Lead.query.filter(

        Lead.lead_level == 'S', Lead.status.in_(['new', 'contacted']),

        Lead.is_opt_out == False, Lead.assigned_to.is_(None)

    ).order_by(Lead.total_score.desc()).limit(10).all()

    return _ok([l.to_dict() for l in s_leads])



@app.route('/api/harvest', methods=['POST'])

def api_harvest():

    if harvest_state['running']:

        return _fail('Harvest already running', 409)

    data = request.get_json()

    areas = data.get('areas', [])

    targets = data.get('targets', [])

    categories = data.get('categories', [])

    if not areas or not targets:

        return _fail('Need at least 1 area and 1 target')

    t = threading.Thread(target=_do_harvest, args=(app, areas, targets, categories), daemon=True)

    t.start()

    return _ok({'status': 'started', 'message': 'Harvest started, poll /api/harvest/progress for updates'})



@app.route('/api/harvest/progress')

def api_harvest_progress():

    return _ok(harvest_state)



@app.route('/api/harvest/channel_stats')

def api_harvest_channel_stats():

    from sqlalchemy import func

    source_map = {

        'baidu_map': '百度地图', '招标平台': '招标平台', 'search': '搜索引擎',

        'classified': '58/百姓网', 'construction': '在建工程', '在建工程': '在建工程', '商品房项目': '商品房项目', 'forum': '问答平台',

    }

    rows = db.session.query(

        Lead.source, func.count(Lead.id),

        func.sum(db.case((Lead.phone.isnot(None) & (Lead.phone != ''), 1), else_=0)),

        func.sum(db.case((Lead.lead_level == 'S', 1), else_=0)),

        func.sum(db.case((Lead.lead_level == 'A', 1), else_=0)),

    ).group_by(Lead.source).all()

    result = []

    for source, total, with_phone, s_count, a_count in rows:

        display = source_map.get(source, source or '未知')

        result.append({'source': source or '', 'display': display, 'total': total,

            'with_phone': with_phone or 0, 's_count': s_count or 0, 'a_count': a_count or 0})

    existing_sources = set(r['source'] for r in result)

    for src, display in source_map.items():

        if src not in existing_sources:

            result.append({'source': src, 'display': display, 'total': 0, 'with_phone': 0, 's_count': 0, 'a_count': 0})

    result.sort(key=lambda x: -x['total'])

    cat_rows = db.session.query(

        Lead.business_category, func.count(Lead.id),

        func.sum(db.case((Lead.phone.isnot(None) & (Lead.phone != ''), 1), else_=0)),

    ).group_by(Lead.business_category).all()

    category_stats = []

    for cat, total, with_phone in cat_rows:

        category_stats.append({'category': cat or '未分类', 'total': total, 'with_phone': with_phone or 0})

    category_stats.sort(key=lambda x: -x['total'])

    return _ok({'channels': result, 'categories': category_stats})



@app.route('/api/harvest/coverage')

def api_harvest_coverage():

    from sqlalchemy import func

    area_stats = db.session.query(

        Lead.area, Lead.customer_type, func.count(Lead.id),

        func.sum(db.case((Lead.phone.isnot(None) & (Lead.phone != ''), 1), else_=0))

    ).group_by(Lead.area, Lead.customer_type).all()

    result = []

    for area, ctype, total, with_phone in area_stats:

        if not area: continue

        result.append({'area': area, 'type': ctype or '', 'total': total, 'with_phone': with_phone or 0})

    result.sort(key=lambda x: (-x['total'], x['area']))

    total_leads = Lead.query.count()

    total_phone = Lead.query.filter(Lead.phone.isnot(None), Lead.phone != '').count()

    areas_covered = len(set(r['area'] for r in result))

    return _ok({'total_leads': total_leads, 'total_with_phone': total_phone,

        'areas_covered': areas_covered, 'details': result})



@app.route('/api/harvest/reset', methods=['POST'])

def api_harvest_reset():

    harvest_state.update({'running': False, 'pct': 0, 'current': '', 'added': 0,

        'with_phone': 0, 'skipped': 0, 'skip_phone': 0, 'skip_name': 0, 'api_calls': 0, 'log_lines': ''})

    return _ok({'status': 'reset'})



@app.route('/api/harvest/projects', methods=['POST'])

def api_harvest_projects():

    if project_harvest_state['running']:

        return _fail('Project harvest already running', 409)

    data = request.get_json()

    source = data.get('source', 'hbcic')

    max_pages = data.get('max_pages', 200)

    area_filter = data.get('area_filter', 'wuhan')

    t = threading.Thread(target=_do_project_harvest, args=(app, source, max_pages, area_filter), daemon=True)

    t.start()

    return _ok({'status': 'started', 'source': source})



@app.route('/api/harvest/projects/progress')

def api_harvest_projects_progress():

    return _ok(project_harvest_state)



@app.route('/api/harvest/projects/reset', methods=['POST'])

def api_harvest_projects_reset():

    project_harvest_state.update({

        'running': False, 'pct': 0, 'current': '', 'added': 0,

        'scanned': 0, 'source': '', 'log_lines': ''

    })

    return _ok({'status': 'reset'})



@app.route('/api/competitors')

def api_competitors():

    """Competitor analysis by region."""

    try:

        import importlib

        import config.competitors as comp_mod

        importlib.reload(comp_mod)

        COMPETITORS = comp_mod.COMPETITORS

        STRENGTH_LABELS = comp_mod.STRENGTH_LABELS

    except (ImportError, AttributeError):

        return _ok({"regions": [], "summary": {}})

    

    from sqlalchemy import func

    # Get our lead counts per area for comparison

    our_leads = dict(db.session.query(

        Lead.area, func.count(Lead.id)

    ).group_by(Lead.area).all())

    

    regions = []

    for area, comps in COMPETITORS.items():

        our_count = our_leads.get(area, 0)

        comp_count = len(comps)

        s_count = sum(1 for c in comps if c.get("strength") == "S")

        a_count = sum(1 for c in comps if c.get("strength") == "A")

        # Opportunity score: more of our leads + fewer competitors = better

        if comp_count == 0:

            opp = "high"

        elif comp_count <= 2 and s_count == 0:

            opp = "medium"

        else:

            opp = "low"

        # Add 企查查 links if not present

        enriched = []

        for c in comps:

            if isinstance(c, dict):

                cc = dict(c)

                if "qcc_url" not in cc:

                    from urllib.parse import quote

                    cc["qcc_url"] = f"https://www.qcc.com/search?key={quote(cc.get('name',''))}"

                    cc["ty_url"] = f"https://www.tianyancha.com/search?key={quote(cc.get('name',''))}"

                enriched.append(cc)

        

        regions.append({

            "area": area,

            "our_leads": our_count,

            "competitor_count": comp_count,

            "strong_competitors": s_count,

            "medium_competitors": a_count,

            "opportunity": opp,

            "competitors": enriched,

        })

    

    # Sort by opportunity (high first)

    opp_order = {"high": 0, "medium": 1, "low": 2}

    regions.sort(key=lambda x: (opp_order.get(x["opportunity"], 3), -x["our_leads"]))

    

    summary = {

        "total_competitors": sum(len(v) for v in COMPETITORS.values()),

        "high_opp_regions": sum(1 for r in regions if r["opportunity"] == "high"),

        "medium_opp_regions": sum(1 for r in regions if r["opportunity"] == "medium"),

        "low_opp_regions": sum(1 for r in regions if r["opportunity"] == "low"),

    }

    return _ok({"regions": regions, "summary": summary})



@app.route('/api/competitors/harvest', methods=['POST'])

def api_competitors_harvest():

    """Harvest competitor data from search engines."""

    from lead_harvester.competitor_harvester import CompetitorHarvester

    from config.competitors import COMPETITORS

    import threading

    

    def _do_harvest():

        try:

            harvester = CompetitorHarvester()

            results = harvester.harvest_all()

            # Save results to competitors config

            for region, companies in results.items():

                if region not in COMPETITORS:

                    COMPETITORS[region] = []

                for c in companies:

                    if not any(existing["name"] == c["name"] for existing in COMPETITORS[region]):

                        COMPETITORS[region].append(c)

            # Write back to file

            _save_competitors(COMPETITORS)

            logger.info(f"[CompetitorHarvest] Done: {sum(len(v) for v in results.values())} new companies")

        except Exception as e:

            logger.error(f"[CompetitorHarvest] Failed: {e}")

    

    threading.Thread(target=_do_harvest, daemon=True).start()

    return _ok({"status": "started", "message": "Competitor harvest started in background"})



def _save_competitors(data):

    """Save competitor data back to config file."""

    import os, json

    cfg_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "config", "competitors.py")

    lines = ['# -*- coding: utf-8 -*-', '"""Auto-generated competitor data. Edit manually or use harvest function."""', '', 'COMPETITORS = {']

    for area, companies in data.items():

        lines.append(f'    "{area}": [')

        for c in companies:

            if isinstance(c, dict):

                lines.append(f'        {json.dumps(c, ensure_ascii=False)},')

            else:

                lines.append(f'        {c},')

        lines.append('    ],')

    lines.append('}')

    lines.append('')

    lines.append('STRENGTH_LABELS = {')

    lines.append('    "S": "强竞争 - 资质全、案例多、价格高",')

    lines.append('    "A": "中竞争 - 有一定实力，价格适中",')

    lines.append('    "B": "弱竞争 - 小型企业或新进入者",')

    lines.append('    "C": "无竞争 - 几乎无同行",')

    lines.append('}')

    _atomic_write_text(cfg_path, "\n".join(lines), encoding="utf-8", syntax_check=False)



@app.route('/api/outreach/feedback')

def api_outreach_feedback():

    """Outreach message feedback summary."""

    from sqlalchemy import func

    from datetime import datetime, timedelta

    

    # Messages by channel and status

    msg_stats = db.session.query(

        Message.channel, Message.status, func.count(Message.id)

    ).group_by(Message.channel, Message.status).all()

    

    by_channel = {}

    for ch, status, count in msg_stats:

        if ch not in by_channel:

            by_channel[ch] = {"total": 0, "sent": 0, "failed": 0, "alert": 0, "replied": 0}

        by_channel[ch]["total"] += count

        if status in by_channel[ch]:

            by_channel[ch][status] += count

        else:

            by_channel[ch][status] = count

    

    # Recent messages with direction

    recent = Message.query.order_by(Message.created_at.desc()).limit(30).all()

    recent_list = []

    for m in recent:

        lead = Lead.query.get(m.lead_id)

        recent_list.append({

            "id": m.id,

            "lead_id": m.lead_id,

            "lead_name": lead.name[:20] if lead else "?",

            "channel": m.channel,

            "direction": m.direction,

            "content": (m.content or "")[:60],

            "status": m.status,

            "template_name": m.template_name or "",

            "created_at": m.created_at.strftime("%m-%d %H:%M") if m.created_at else "",

        })

    

    # Incoming messages (replies)

    replies = Message.query.filter_by(direction="in").order_by(Message.created_at.desc()).limit(20).all()

    reply_list = []

    for m in replies:

        lead = Lead.query.get(m.lead_id)

        reply_list.append({

            "id": m.id,

            "lead_id": m.lead_id,

            "lead_name": lead.name[:20] if lead else "?",

            "channel": m.channel,

            "content": (m.content or "")[:80],

            "status": m.status,

            "created_at": m.created_at.strftime("%m-%d %H:%M") if m.created_at else "",

        })

    

    return _ok({

        "by_channel": by_channel,

        "recent": recent_list,

        "replies": reply_list,

    })



@app.route('/api/leads', methods=['GET'])

def api_get_leads():

    page = request.args.get('page', 1, type=int)

    per_page = request.args.get('per_page', 50, type=int)

    status = request.args.get('status')

    q = Lead.query

    if status: q = q.filter_by(status=status)

    pagination = q.order_by(Lead.created_at.desc()).paginate(page=page, per_page=per_page, error_out=False)

    try:
        return _ok({'leads': [l.to_dict() for l in pagination.items], 'total': pagination.total, 'page': page, 'pages': pagination.pages})
    except Exception as e:
        return _fail(str(e), 500)



@app.route('/api/leads', methods=['POST'])

def api_add_lead():

    data = request.get_json()

    if not data: return _fail('Invalid')

    from crm.scoring import score_lead

    lead = add_lead(data)

    score_lead(lead)

    db.session.commit()

    return _ok(lead.to_dict()), 201



@app.route('/api/leads/<int:lid>', methods=['PUT'])

def api_update_lead(lid):

    data = request.get_json()

    lead = Lead.query.get_or_404(lid)

    for k, v in data.items():

        if hasattr(lead, k) and k not in ('id', 'created_at'): setattr(lead, k, v)

    lead.updated_at = datetime.now()

    db.session.commit()

    return _ok(lead.to_dict())



@app.route('/api/leads/<int:lid>/status', methods=['POST'])

def api_update_status(lid):

    data = request.get_json()

    lead = update_lead_status(lid, data.get('status'), data.get('note'))

    if lead: return _ok(lead.to_dict())

    return _fail('Not found', 404)



@app.route('/api/leads/<int:lid>/score', methods=['POST'])

def api_update_score(lid):

    data = request.get_json()

    lead = update_lead_score(lid, data.get('delta', 0))

    if lead: return _ok(lead.to_dict())

    return _fail('Not found', 404)



@app.route('/api/leads/<int:lid>/set_level', methods=['POST'])

def api_set_level(lid):

    """Set lead level directly (for batch operations)."""

    data = request.get_json()

    lead = Lead.query.get_or_404(lid)

    level = data.get('set_level', '')

    if level in ('S', 'A', 'B', 'C'):

        lead.lead_level = level

        lead.updated_at = datetime.now()

        db.session.commit()

        return _ok(lead.to_dict())

    return _fail('Invalid level')



@app.route('/api/leads/<int:lid>/message', methods=['POST'])

def api_send_message(lid):

    data = request.get_json()

    msg = record_message(lid, data.get('channel', 'manual'), 'out', data.get('content', ''), status='sent')

    return _ok(msg.to_dict())



@app.route('/api/handoff-leads')

def api_handoff():

    return _ok([l.to_dict() for l in get_human_handoff_leads()])



@app.route('/api/tasks')

def api_tasks():

    return _ok([l.to_dict() for l in TaskLog.query.order_by(TaskLog.started_at.desc()).limit(20).all()])



@app.route('/api/import', methods=['POST'])

def api_import():

    data = request.get_json()

    if not isinstance(data, list): return _fail('Need JSON array')

    from crm.database import add_leads_batch

    from crm.scoring import score_lead

    added, skipped = add_leads_batch(data)

    for lead in Lead.query.filter(Lead.total_score == 0).all():

        score_lead(lead)

    db.session.commit()

    return _ok({'added': added, 'skipped': skipped})



@app.route('/api/parse_bid/<int:lid>')

def api_parse_bid(lid):

    lead = Lead.query.get_or_404(lid)

    if not lead.source_url:

        return _fail('No source URL')

    from crm.bid_parser import parse_bid_page, format_bid_info

    info = parse_bid_page(lead.source_url)

    return _ok(info)



@app.route('/api/parse_bid/<int:lid>/save', methods=['POST'])

def api_save_bid_info(lid):

    lead = Lead.query.get_or_404(lid)

    data = request.get_json()

    if not data:

        return _fail('No data')

    from crm.bid_parser import format_bid_info

    formatted = format_bid_info(data)

    lead.notes = (lead.notes or "") + f"\n[\u62db\u6807\u4fe1\u606f {datetime.now().strftime('%m-%d')}]\n{formatted}"

    if data.get("budget"):

        lead.bid_budget_text = data["budget"]

    if data.get("budget_num"):

        lead.bid_budget = data["budget_num"]

    if data.get("bid_deadline"):

        lead.bid_deadline = data["bid_deadline"]

    if data.get("open_time"):

        lead.bid_open_time = data["open_time"]

    if data.get("purchaser"):

        lead.bid_purchaser = data["purchaser"]

    if data.get("agency"):

        lead.bid_agency = data["agency"]

    phone = data.get("purchaser_contact", "")

    phone_match = re.search(r"1[3-9]\d{9}", phone)

    if phone_match and not lead.phone:

        lead.phone = phone_match.group()

    db.session.commit()

    return _ok(lead.to_dict())

def _get_followup_script(lead):
    """Generate follow-up script based on customer type, level, and category."""
    ctype = (lead.customer_type or "").strip()
    level = lead.lead_level or "C"
    cat = (lead.business_category or "").strip()
    area = lead.area or "本地"
    name = lead.name or "贵公司"

    product = "车棚/遮阳棚"
    if "光伏" in cat:
        product = "光伏车棚"
    elif "玻璃" in cat:
        product = "玻璃遮阳棚"

    diff = "我们建筑一级资质，膜布+钢结构自主生产，1000+案例，24h出初稿"

    by_type = {
        "4S店": "电话开场白: 您好，我是XX膜结构的，看到您这边是{area}的4S店。最近我们帮附近几家4S店做了门头车棚升级，到店体验提升明显。{diff}，可以免费上门看现场出方案。您看什么时候方便？\n微信跟进: 发案例图片，这是我们在{area}做的4S店车棚项目，有空我们上门量场出图。",
        "工厂": "电话开场白: 您好，我是XX膜结构的。看到您这边是{area}的工厂，想问一下厂区车棚有没有在考虑升级？最近不少工厂都在换车棚，雨季不用担心漏水。{diff}，可以免费上门测量出方案。\n微信跟进: 发厂区车棚案例图，同行工厂实拍，钢结构+膜布材料耐用且维护低。您这边大概多大面积？我先出个方案。",
        "物业": "电话开场白: 您好，我是XX膜结构的。看到您这边是{area}的小区物业，想了解一下车棚或充电桩雨棚有没有需求？附近几个小区今年刚装好，业主反馈很好。{diff}，24h出初稿，3天出终稿。\n微信跟进: 发小区车棚案例图，同类型小区完工照片，业主群反响不错。如果有计划，我们免费上门量场出图报价。",
        "招标项目": "电话开场白: 您好，我是XX膜结构的，看到贵司有个{product}项目在招标。我们建筑一级资质，1000+工程案例，可配合招标全流程，24h出初稿。我发一份资质和案例给您参考？\n微信跟进: 发资质证书+案例集，可继续深入沟通，我们可配合出方案、报价、资质文件全套。",
    }

    default_tmpl = "电话开场白: 您好，我是XX膜结构的。看到您这边是{area}的，想了解一下{product}有没有需求？{diff}，可以免费上门测量出方案。\n微信跟进: 发案例图片，这是我们在{area}做的项目，有空我们上门量场免费出图。"

    tmpl = by_type.get(ctype, default_tmpl)
    script = tmpl.format(area=area, diff=diff, product=product, name=name)

    prefix_map = {
        "S": "[S级-立即电话] ",
        "A": "[A级-24h内联系] ",
        "B": "[B级-短信/微信触达] ",
        "C": "[C级-有空再跟进] ",
    }
    return prefix_map.get(level, "[C级] ") + script


@app.route('/api/leads/callable')
def api_callable_leads():
    """获取可拨打的高质量线索"""
    from lead_harvester.phone_verifier import PhoneVerifier

    min_score = request.args.get('min_score', 50, type=int)
    limit = request.args.get('limit', 100, type=int)

    callable_leads = PhoneVerifier.filter_callable_leads(min_score=min_score)

    return _ok({
        "total": len(callable_leads),
        "leads": [{
            "id": item["lead"].id,
            "name": item["lead"].name,
            "phone": item["lead"].phone,
            "contact_person": item["lead"].contact_person,
            "phone_score": item["phone_score"],
            "has_contact": item["has_contact"],
            "source": item["lead"].source,
            "area": item["lead"].area,
            "customer_type": item["lead"].customer_type,
        } for item in callable_leads[:limit]]
    })


@app.route('/api/leads/export')


def api_export_leads_excel():
    """导出线索为 Excel（修复函数名重复导致的路由影子问题）."""
    from flask import send_file
    import io, openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    # Apply same filters as leads_page
    status = request.args.get('status', '')
    source = request.args.get('source', '')
    business_category = request.args.get('business_category', '')
    customer_type = request.args.get('customer_type', '')
    area = request.args.get('area', '')
    has_phone = request.args.get('has_phone', '')
    lead_level = request.args.get('lead_level', '')

    q = Lead.query
    if status: q = q.filter_by(status=status)
    if source: q = q.filter_by(source=source)
    if business_category: q = q.filter_by(business_category=business_category)
    if customer_type: q = q.filter_by(customer_type=customer_type)
    if area: q = q.filter(Lead.area.like(f'%{area}%'))
    if has_phone == 'yes': q = q.filter(Lead.phone.isnot(None), Lead.phone != '')
    elif has_phone == 'no': q = q.filter((Lead.phone.is_(None)) | (Lead.phone == ''))
    if lead_level: q = q.filter_by(lead_level=lead_level)
    only_new = request.args.get('only_new', '')
    if only_new == 'yes':
        q = q.filter(Lead.exported_at.is_(None))

    leads = q.order_by(Lead.lead_level.asc(), Lead.total_score.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "线索列表"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="001529", end_color="001529", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )

    headers = ["等级", "名称", "联系人", "电话", "微信", "类型", "区域", "业务分类", "状态", "总分", "需求描述", "备注", "跟进话术", "来源", "创建时间", "上次导出时间", "是否新导出"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Level color fills
    level_fills = {
        "S": PatternFill(start_color="FFF1F0", end_color="FFF1F0", fill_type="solid"),
        "A": PatternFill(start_color="FFF7E6", end_color="FFF7E6", fill_type="solid"),
        "B": PatternFill(start_color="E6F7FF", end_color="E6F7FF", fill_type="solid"),
        "C": PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
    }

    status_labels = {"new": "新线索", "contacted": "已联系", "interested": "有意向", "quoting": "报价中", "closed_won": "已成交", "closed_lost": "已流失"}

    for row_idx, lead in enumerate(leads, 2):
        row_data = [
            lead.lead_level or "C",
            lead.name or "",
            lead.contact_person or "",
            lead.phone or "",
            lead.wechat or "",
            lead.customer_type or "",
            lead.area or "",
            lead.business_category or "",
            status_labels.get(lead.status, lead.status),
            lead.total_score or 0,
            (lead.demand_desc or "")[:200],
            (lead.notes or "")[:200],
            _get_followup_script(lead),
            lead.source or "",
            lead.created_at.strftime("%Y-%m-%d %H:%M") if lead.created_at else "",
            lead.exported_at.strftime("%Y-%m-%d %H:%M") if lead.exported_at else "",
            "新" if not lead.exported_at else "已导出",
        ]
        fill = level_fills.get(lead.lead_level)
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if fill:
                cell.fill = fill

    # Auto column widths
    col_widths = [6, 30, 10, 15, 15, 10, 10, 12, 8, 6, 40, 40, 50, 12, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto filter
    ws.auto_filter.ref = ws.dimensions

    # Mark exported leads
    now = datetime.now()
    for lead in leads:
        lead.exported_at = now
    db.session.commit()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"leads_export_{len(leads)}_条.xlsx"
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)

@app.route('/api/leads/<int:lid>/contacts')



@app.route('/api/notify/test', methods=['POST'])
def api_test_notify():
    """测试通知发送"""
    from auto_outreach.notifications import send_notification
    ok = send_notification("测试通知", "数字营销系统通知功能正常")
    return _ok({"sent": ok})


@app.route('/api/notify/daily_report', methods=['POST'])
def api_daily_report():
    """手动触发每日报告"""
    from auto_outreach.notifications import send_daily_report
    msg = send_daily_report()
    return _ok({"sent": True, "message": msg})


@app.route('/api/notify/config')
def api_notify_config():
    """查看通知配置状态"""
    from config import settings
    return _ok({
        "telegram_configured": bool(settings.TELEGRAM_BOT_TOKEN and settings.TELEGRAM_CHAT_ID),
        "wechat_configured": bool(settings.WECHAT_CORP_ID and settings.WECHAT_CORP_SECRET),
        "serverchan_configured": bool(settings.SERVERCHAN_KEY),
        "s_level_instant": settings.S_LEVEL_INSTANT_NOTIFY,
        "daily_report_hour": settings.DAILY_REPORT_HOUR,
        "daily_report_minute": settings.DAILY_REPORT_MINUTE,
    })

def api_lead_project_contacts(lid):
    """Extract project contacts from related construction project leads."""
    import re as _re
    lead = Lead.query.get_or_404(lid)

    # Find related project leads by area or name similarity
    related = []
    if lead.area:
        related = Lead.query.filter(
            Lead.id != lid,
            Lead.area.like(f'%{lead.area}%'),
        ).limit(30).all()

    # Also find by name keyword overlap
    if lead.name:
        for part in [lead.name[:4], lead.name[2:6]]:
            if len(part) >= 2:
                more = Lead.query.filter(Lead.id != lid, Lead.name.like(f'%{part}%')).limit(10).all()
                for m in more:
                    if m.id not in [r.id for r in related]:
                        related.append(m)

    # Deduplicate
    seen = set()
    unique_related = []
    for r in related:
        if r.id not in seen:
            seen.add(r.id)
            unique_related.append(r)

    # Extract contact info with aggressive phone extraction
    phone_re = _re.compile(r'1[3-9]\d{9}')
    landline_re = _re.compile(r'0\d{2,3}[- ]?\d{7,8}')
    contacts = []
    for r in unique_related:
        info = {"id": r.id, "name": r.name, "source": r.source}
        all_text = f"{r.phone or ''} {r.contact_person or ''} {r.notes or ''} {r.demand_desc or ''} {r.address or ''}"
        phones = phone_re.findall(all_text)
        landlines = landline_re.findall(all_text)
        if r.phone and r.phone not in phones:
            phones.insert(0, r.phone)
        phones = list(dict.fromkeys(phones))
        landlines = list(dict.fromkeys(landlines))
        info["phones"] = phones
        info["landlines"] = landlines
        # Extract contact person from notes if not set
        contact = r.contact_person or ""
        if not contact:
            cp_match = _re.search(r'(?:\u8054\u7cfb\u4eba|\u8d1f\u8d23\u4eba|\u9879\u76ee\u7ecf\u7406)[\uff1a: ]*([\u4e00-\u9fa5]{2,6})', all_text)
            if cp_match:
                contact = cp_match.group(1)
        info["contact_person"] = contact
        info["address"] = r.address or ""
        info["area"] = r.area or ""
        info["customer_type"] = r.customer_type or ""
        info["notes_snippet"] = (r.notes or "")[:200]
        info["demand_snippet"] = (r.demand_desc or "")[:200]
        contacts.append(info)

    # Sort: contacts with phones first, then by name length (shorter = more likely real company)
    contacts.sort(key=lambda c: (0 if c["phones"] or c["landlines"] else 1, len(c["name"])))

    return _ok({"lead_id": lid, "total": len(contacts), "contacts": contacts})
