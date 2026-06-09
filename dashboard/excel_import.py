# -*- coding: utf-8 -*-
"""Excel import API for Tianyancha/enterprise data"""
import os
import re
from datetime import datetime
from flask import request, jsonify
from loguru import logger

from crm.models import db, Lead
from crm.scoring import score_lead


# 天眼查/企查查列名 -> Lead 字段映射
COLUMN_MAP = {
    # 天眼查标准导出列
    "企业名称": "name",
    "公司名称": "name",
    "名称": "name",
    "法定代表人": "contact_person",
    "法人": "contact_person",
    "法人代表": "contact_person",
    "联系人": "contact_person",
    "联系电话": "phone",
    "电话": "phone",
    "手机号": "phone",
    "手机": "phone",
    "联系方式": "phone",
    "企业地址": "address",
    "地址": "address",
    "注册地址": "address",
    "经营范围": "business_scope",
    "工商信息": "business_scope",
    "统一社会信用代码": "credit_code",
    "邮箱": "email",
    "电子邮箱": "email",
    "微信": "wechat",
    "备注": "notes",
    "来源": "source_hint",
    # 英文列名（有些导出工具）
    "Company Name": "name",
    "Legal Representative": "contact_person",
    "Phone": "phone",
    "Address": "address",
    "Email": "email",
}

# 电话号码正则
PHONE_RE = re.compile(r"1[3-9]\d{9}")
LANDLINE_RE = re.compile(r"0\d{2,3}[-]?\d{7,8}")


def parse_excel_import(file_storage, default_source="import", default_area="武汉"):
    """解析上传的 Excel 文件，返回标准化的线索数据列表

    Args:
        file_storage: Flask FileStorage 对象
        default_source: 默认来源标签
        default_area: 默认区域

    Returns:
        list of dict: 标准化后的线索数据
    """
    import openpyxl

    try:
        wb = openpyxl.load_workbook(file_storage, read_only=True, data_only=True)
    except Exception as e:
        logger.error(f"Excel解析失败: {e}")
        return []

    ws = wb.active
    if not ws:
        return []

    # 读取表头，建立列映射
    headers = []
    col_map = {}  # col_index -> lead_field
    for col_idx, cell in enumerate(ws[1]):
        header = str(cell.value).strip() if cell.value else ""
        headers.append(header)
        # 尝试匹配已知列名
        for known_name, field in COLUMN_MAP.items():
            if known_name in header or header in known_name:
                col_map[col_idx] = field
                break

    # 如果没有找到任何匹配的列，尝试按位置猜测
    if not col_map:
        col_map = _guess_columns_by_position(headers)

    # 解析数据行
    leads = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(v is None for v in row):
            continue

        data = {}
        for col_idx, value in enumerate(row):
            if col_idx in col_map and value is not None:
                field = col_map[col_idx]
                data[field] = str(value).strip()

        # 提取电话号码
        phone = data.get("phone", "")
        if not phone:
            # 从其他字段尝试提取
            for field in ["address", "business_scope", "notes"]:
                if field in data:
                    match = PHONE_RE.search(data[field])
                    if match:
                        phone = match.group()
                        break
                    match = LANDLINE_RE.search(data[field])
                    if match:
                        phone = match.group()
                        break

        # 标准化电话号码
        phone = _normalize_phone(phone)

        # 企业名称
        name = data.get("name", "")
        if not name or len(name) < 2:
            continue

        # 跳过竞争对手
        if _is_competitor(name):
            continue

        # 构建线索数据
        lead_data = {
            "name": name[:100],
            "contact_person": data.get("contact_person", ""),
            "phone": phone,
            "email": data.get("email", ""),
            "wechat": data.get("wechat", ""),
            "address": data.get("address", "")[:300] if data.get("address") else "",
            "source": default_source,
            "area": _detect_area(name, data.get("address", ""), default_area),
            "customer_type": _detect_customer_type(name, data.get("business_scope", "")),
            "demand_desc": data.get("business_scope", "")[:200] if data.get("business_scope") else "",
            "notes": data.get("notes", ""),
        }
        leads.append(lead_data)

    wb.close()
    return leads


def _guess_columns_by_position(headers):
    """按位置猜测列映射（兜底方案）"""
    col_map = {}
    # 常见位置：0=名称, 1=法人, 2=电话, 3=地址
    position_guesses = [
        (0, "name"),
        (1, "contact_person"),
        (2, "phone"),
        (3, "address"),
        (4, "email"),
    ]
    for idx, field in position_guesses:
        if idx < len(headers):
            col_map[idx] = field
    return col_map


def _normalize_phone(phone: str) -> str:
    """标准化电话号码"""
    if not phone:
        return ""
    # 去掉空格和横杠
    phone = phone.replace(" ", "").replace("-", "").replace("(", "").replace(")", "")
    # 提取手机号
    match = PHONE_RE.search(phone)
    if match:
        return match.group()
    # 提取座机号
    match = LANDLINE_RE.search(phone)
    if match:
        return match.group()
    return phone[:20]


def _is_competitor(name: str) -> bool:
    """检测是否为竞争对手"""
    competitor_kws = [
        "膜结构", "车棚厂", "遮阳棚", "雨棚厂", "张拉膜",
        "膜结构公司", "膜结构工程", "钢结构车棚",
    ]
    for kw in competitor_kws:
        if kw in name:
            return True
    return False


def _detect_area(name: str, address: str, default: str) -> str:
    """从名称或地址中检测区域"""
    text = f"{name} {address}"
    areas = ["武汉", "黄石", "鄂州", "孝感", "黄冈", "咸宁", "荆州", "荆门",
             "宜昌", "襄阳", "十堰", "随州", "恩施", "仙桃", "潜江", "天门",
             "长沙", "岳阳", "常德", "南昌", "九江", "信阳", "南阳"]
    for area in areas:
        if area in text:
            return area
    return default


def _detect_customer_type(name: str, scope: str) -> str:
    """从企业名称和经营范围推断客户类型"""
    text = f"{name} {scope}"
    type_rules = [
        (["物业", "物业管理", "物业服务"], "物业"),
        (["4S", "汽车销售", "汽车贸易", "汽车服务"], "4S店"),
        (["学校", "学院", "大学", "中学", "小学", "幼儿园"], "学校"),
        (["医院", "卫生院", "诊所", "医疗"], "医院"),
        (["工厂", "制造", "生产", "加工", "工业"], "工厂"),
        (["新能源", "充电", "光伏"], "新能源"),
        (["酒店", "宾馆", "饭店"], "酒店"),
        (["商场", "购物", "百货", "商贸"], "商业"),
        (["房地产", "置业", "地产"], "物业"),  # 房地产开发商也归为物业类
    ]
    for keywords, ctype in type_rules:
        for kw in keywords:
            if kw in text:
                return ctype
    return "其他"


def save_imported_leads(leads_data: list) -> dict:
    """保存导入的线索到数据库

    Returns:
        dict: {"added": int, "skipped": int, "details": list}
    """
    added = 0
    skipped = 0
    details = []

    for data in leads_data:
        try:
            # 电话去重
            if data.get("phone"):
                existing = Lead.query.filter_by(phone=data["phone"]).first()
                if existing:
                    skipped += 1
                    continue

            # 名称去重
            if data.get("name"):
                existing = Lead.query.filter_by(name=data["name"]).first()
                if existing:
                    skipped += 1
                    continue

            # 创建线索
            lead = Lead(
                name=data.get("name", ""),
                contact_person=data.get("contact_person", ""),
                phone=data.get("phone", ""),
                email=data.get("email", ""),
                wechat=data.get("wechat", ""),
                address=data.get("address", ""),
                source=data.get("source", "import"),
                area=data.get("area", ""),
                customer_type=data.get("customer_type", ""),
                demand_desc=data.get("demand_desc", ""),
                notes=data.get("notes", ""),
            )
            db.session.add(lead)
            db.session.flush()  # 获取ID

            # 评分
            score_lead(lead)
            added += 1

        except Exception as e:
            logger.debug(f"导入线索失败: {e}")
            skipped += 1

    db.session.commit()

    return {
        "added": added,
        "skipped": skipped,
        "total": len(leads_data),
    }
