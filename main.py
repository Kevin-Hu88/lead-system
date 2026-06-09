# -*- coding: utf-8 -*-
"""
Main entry - with integrated scoring, tiered outreach, and sleep mechanism
"""
import os, sys, argparse, threading, time
from datetime import datetime
from loguru import logger

logger.remove()
logger.add(sys.stderr, level="INFO", format="{time:HH:mm:ss} | {level: <7} | {message}")

# 只在本地环境写日志文件
if not os.getenv("RENDER"):
    logger.add("logs/app_{time:YYYY-MM-DD}.log", rotation="1 day", retention="30 days", level="DEBUG")

# 导入 app 供 gunicorn 使用
from dashboard.app import app


def run_scheduler(app):
    import time as t
    from apscheduler.schedulers.background import BackgroundScheduler
    from lead_harvester.harvester_main import run_harvesters
    from auto_outreach.tiered_outreach import auto_contact_by_level
    from crm.scoring import score_all_leads, mark_sleep_leads

    scheduler = BackgroundScheduler(timezone="Asia/Shanghai")

    scheduler.add_job(run_harvesters, 'interval', hours=2, args=[app],
                      id='harvest', name='线索采集', replace_existing=True)
    scheduler.add_job(score_all_leads, 'interval', hours=1, args=[app],
                      id='scoring', name='自动打分', replace_existing=True)
    scheduler.add_job(auto_contact_by_level, 'interval', minutes=30, args=[app],
                      id='outreach', name='分级触达', replace_existing=True)
    scheduler.add_job(mark_sleep_leads, 'cron', hour=2, args=[app],
                      id='sleep', name='沉睡标记', replace_existing=True)

    scheduler.start()
    logger.info("Scheduler started: harvest(2h), scoring(1h), outreach(30m), sleep(2am)")
    try:
        while True: t.sleep(60)
    except (KeyboardInterrupt, SystemExit):
        scheduler.shutdown()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--web-only", action="store_true")
    parser.add_argument("--harvest", action="store_true")
    parser.add_argument("--outreach", action="store_true", help="Run tiered outreach once")
    parser.add_argument("--score", action="store_true", help="Re-score all leads")
    parser.add_argument("--import", dest="import_file", help="Import Excel")
    parser.add_argument("--stats", action="store_true")
    parser.add_argument("--port", type=int, default=5000)
    args = parser.parse_args()

    if args.import_file:
        if not os.path.exists(args.import_file):
            logger.error(f"File not found: {args.import_file}"); sys.exit(1)
        from openpyxl import load_workbook
        from crm.database import add_lead
        from crm.scoring import score_lead
        wb = load_workbook(args.import_file, read_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        field_map = {
            "名称": "name", "公司": "name", "客户名称": "name",
            "联系人": "contact_person", "姓名": "contact_person",
            "电话": "phone", "手机": "phone",
            "邮箱": "email", "微信": "wechat", "地址": "address",
            "类型": "customer_type", "区域": "area",
            "需求": "demand_desc", "产品": "product_interest", "备注": "notes",
        }
        fields = [field_map.get((h or "").strip()) for h in headers]
        added = 0
        with app.app_context():
            for row in ws.iter_rows(min_row=2, values_only=True):
                data = {"source": "import"}
                for i, val in enumerate(row):
                    if val and i < len(fields) and fields[i]:
                        data[fields[i]] = str(val).strip()
                if not data.get("name"): continue
                try:
                    lead = add_lead(data)
                    score_lead(lead)
                    from config.scoring import AUTO_CATEGORY_RULES
                    text = f"{data.get('product_interest', '')} {data.get('demand_desc', '')} {data.get('customer_type', '')}".lower()
                    for kw, (primary, secondary) in AUTO_CATEGORY_RULES.items():
                        if kw in text:
                            lead.business_category = primary
                            break
                    if lead.created_at and (datetime.utcnow() - lead.created_at).seconds < 60:
                        added += 1
                except: pass
            from crm.models import db; db.session.commit()
        wb.close()
        logger.info(f"Imported {added} leads")
        return

    if args.stats:
        from crm.database import get_dashboard_stats
        from crm.models import Lead
        with app.app_context():
            stats = get_dashboard_stats()
            levels = {l: Lead.query.filter_by(lead_level=l).count() for l in ['S','A','B','C']}
            optout = Lead.query.filter_by(is_opt_out=True).count()
            sleep = Lead.query.filter_by(sleep_status=1).count()
            print()
            print("=" * 50)
            print("  Lead Generation System - Stats")
            print("=" * 50)
            print(f"  Total:       {stats['total_leads']}")
            print(f"  Today:       {stats['new_today']}")
            print(f"  This week:   {stats['week_leads']}")
            print(f"  With phone:  {Lead.query.filter(Lead.phone.isnot(None), Lead.phone!='').count()}")
            print("-" * 50)
            print(f"  S (high):    {levels['S']}")
            print(f"  A (medium):  {levels['A']}")
            print(f"  B (normal):  {levels['B']}")
            print(f"  C (low):     {levels['C']}")
            print(f"  Opt-out:     {optout}")
            print(f"  Sleeping:    {sleep}")
            print("-" * 50)
            callable_count = Lead.query.filter(
                Lead.phone.isnot(None), Lead.phone != '',
                Lead.phone_score >= 50, Lead.is_opt_out == False
            ).count()
            high_quality = Lead.query.filter(
                Lead.phone.isnot(None), Lead.phone != '',
                Lead.phone_score >= 70, Lead.is_opt_out == False
            ).count()
            print(f"  Callable (>=50): {callable_count}")
            print(f"  High quality(>=70): {high_quality}")
            print("=" * 50)
        return

    if args.harvest:
        from lead_harvester.harvester_main import run_harvesters
        run_harvesters(app)
        return

    if args.outreach:
        from auto_outreach.tiered_outreach import auto_contact_by_level
        auto_contact_by_level(app)
        return

    if args.score:
        from crm.scoring import score_all_leads
        score_all_leads(app)
        return

    if not args.web_only:
        t_thread = threading.Thread(target=run_scheduler, args=(app,), daemon=True)
        t_thread.start()

    logger.info(f"Web dashboard: http://localhost:{args.port}")
    app.run(host="0.0.0.0", port=args.port, debug=False, use_reloader=False)


if __name__ == "__main__":
    main()